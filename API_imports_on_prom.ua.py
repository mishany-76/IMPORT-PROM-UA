import os
import io
import gzip
import boto3
import requests
import json
import time  # ДОБАВЛЕНО: модуль для создания пауз
from openpyxl import load_workbook
from google.oauth2.service_account import Credentials
from googleapiclient.discovery import build
from googleapiclient.http import MediaIoBaseDownload

# --- НАСТРОЙКИ ---
# Подгружаем данные из переменных окружения (для GitHub Actions)
# или берем дефолтные значения для локальной проверки
R2_ACCESS_KEY = os.getenv('R2_ACCESS_KEY', 'R2_ACCESS_KEY')
R2_SECRET_KEY = os.getenv('R2_SECRET_KEY', 'R2_SECRET_KEY')
R2_ACCOUNT_ID = os.getenv('R2_ACCOUNT_ID', 'R2_ACCOUNT_ID')
R2_BUCKET_NAME = os.getenv('R2_BUCKET_NAME', 'R2_BUCKET_NAME')
R2_ENDPOINT = f'https://{R2_ACCOUNT_ID}.r2.cloudflarestorage.com'
# Домен, который привязан к бакету (Custom Domain или r2.dev)
R2_PUBLIC_DOMAIN = os.getenv('R2_PUBLIC_DOMAIN', 'R2_PUBLIC_DOMAIN')

# ДОБАВЛЕНО: Настройки пауз в секундах
DELAY_BEFORE_FIRST = int(os.getenv('DELAY_BEFORE_FIRST', 10))
DELAY_BETWEEN_REQUESTS = int(os.getenv('DELAY_BETWEEN_REQUESTS', 20))

FILE_NAME = 'import_data.xlsx.gz'
PUBLIC_URL = f'{R2_PUBLIC_DOMAIN}/{FILE_NAME}'

# Список токенов для магазинов
PROM_TOKENS = [
    os.getenv('PROM_TOKEN_City', 'PROM_TOKEN_City'),
    os.getenv('PROM_TOKEN_Best', 'PROM_TOKEN_Best'),
    os.getenv('PROM_TOKEN_Vaillant', 'PROM_TOKEN_Vaillant'),
]

SPREADSHEET_ID = os.getenv('IMPORT_PROM_UA_SPREADSHEET_ID', 'IMPORT_PROM_UA_SPREADSHEET_ID')
SERVICE_ACCOUNT_FILE = 'key_sheet.json'


def get_gspread_xlsx():
    """Скачивает Google таблицу в формате XLSX в память"""
    print("1. Авторизация в Google и скачивание XLSX...")
    creds = Credentials.from_service_account_file(
        SERVICE_ACCOUNT_FILE,
        scopes=['https://www.googleapis.com/auth/drive.readonly']
    )
    service = build('drive', 'v3', credentials=creds)

    request = service.files().export_media(
        fileId=SPREADSHEET_ID,
        mimeType='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )

    fh = io.BytesIO()
    downloader = MediaIoBaseDownload(fh, request)
    done = False
    while done is False:
        status, done = downloader.next_chunk()
    return fh.getvalue()


def process_xlsx(xlsx_content):
    """Удаляет лист 'Systemic' и сжимает в Gzip"""
    print("2. Обработка файла: удаление листа 'Systemic'...")
    wb = load_workbook(io.BytesIO(xlsx_content))

    if "Systemic" in wb.sheetnames:
        del wb["Systemic"]
        print("   Лист 'Systemic' удален.")

    # Сохраняем обработанный XLSX в буфер
    processed_xlsx_bin = io.BytesIO()
    wb.save(processed_xlsx_bin)

    print("3. Сжатие файла в Gzip...")
    gzip_buffer = io.BytesIO()
    with gzip.GzipFile(fileobj=gzip_buffer, mode='wb') as f:
        f.write(processed_xlsx_bin.getvalue())

    return gzip_buffer.getvalue()


def upload_to_r2(data):
    """Загрузка файла в Cloudflare R2"""
    print(f"4. Загрузка в Cloudflare R2 (Бакет: {R2_BUCKET_NAME})...")
    s3 = boto3.client(
        's3',
        endpoint_url=R2_ENDPOINT,
        aws_access_key_id=R2_ACCESS_KEY,
        aws_secret_access_key=R2_SECRET_KEY,
        region_name='auto'
    )

    s3.put_object(
        Bucket=R2_BUCKET_NAME,
        Key=FILE_NAME,
        Body=data,
        ContentType='application/x-gzip',
        ACL='public-read'  # Если бакет требует явного указания, иначе настраивается в панели R2
    )
    print(f"   Файл доступен по ссылке: {PUBLIC_URL}")


def trigger_prom_import():
    """Запуск импорта с полным маппингом твоих заголовков"""
    print("5. Запуск импорта через Prom API...")
    import_url = "https://my.prom.ua/api/v1/products/import_url"

    # Мы перечисляем категории полей.
    # Каждая категория (например 'name') охватывает и рус, и укр колонки в твоем файле.
    payload = {
        "url": PUBLIC_URL,
        "force_update": True,
        "only_available": True,  # «Завантажити позиції "В наявності"»
        "only_update": True,
        "mark_missing_product_as": "deleted",
        "updated_fields": [
            "name",  # Назва_позиції + Назва_позиції_укр
            "name_ua",
            "description",  # Опис + Опис_укр
            "description_ua",
            "keywords",  # Пошукові_запити + Пошукові_запити_укр
            "keywords_ua",
            "sku",  # Код_товару
            "price",  # Ціна + Оптова_ціна
            "old_price",
            "images",  # Посилання_зображення
            "photo",  # (Дублируем для надежности под UI)
            "presence",  # Наявність
            "quantity",  # Кількість
            "group",  # Номер_групи + Назва_групи
            "discount",  # Знижка
            "notes"  # Особисті нотатки
        ]
    }

    is_first = True  # ДОБАВЛЕНО: Флаг для проверки первого запроса

    for token in PROM_TOKENS:
        if not token or token == 'токен_1': continue

        # ДОБАВЛЕНО: Логика создания пауз
        if is_first:
            print(f"   Ожидание {DELAY_BEFORE_FIRST} сек. перед загрузкой первого магазина...")
            time.sleep(DELAY_BEFORE_FIRST)
            is_first = False
        else:
            print(f"   Ожидание {DELAY_BETWEEN_REQUESTS} сек. перед загрузкой следующего магазина...")
            time.sleep(DELAY_BETWEEN_REQUESTS)

        headers = {
            'Authorization': f'Bearer {token}',
            'Content-Type': 'application/json',
            'X-Language': 'uk'  # ГОВОРИМ ПРОМУ, ЧТО МЫ ПРИШЛИ С УКРАИНСКИМ КОНТЕКСТОМ
        }

        try:
            response = requests.post(import_url, json=payload, headers=headers)
            if response.status_code == 200:
                data = response.json()
                print(f"   [OK] Магазин {token[:10]}... Импорт запущен. ID: {data.get('id')}")
            else:
                print(f"   [Error] {token[:10]}: {response.status_code} - {response.text}")
        except Exception as e:
            print(f"   [Exception] Ошибка при запросе к Prom: {e}")


if __name__ == "__main__":
    try:
        # Логика "на лету"
        raw_xlsx = get_gspread_xlsx()
        gzipped_xlsx = process_xlsx(raw_xlsx)
        upload_to_r2(gzipped_xlsx)
        trigger_prom_import()
        print("\nГотово! Все процессы завершены.")
    except Exception as e:
        print(f"\nКритическая ошибка: {e}")
